"""
Exporta os palpites dos 3 boloes do Controle_Boloes_Copa2026.xlsx para um CSV
plano que o job de email consome (sem depender de openpyxl em runtime).

Saida: bolao_polla/palpites.csv com colunas:
    grupo, jogo, home, away, zap_h, zap_a, cli_h, cli_a, polla_h, polla_a

home/away ficam com o nome em INGLES (casavel com a API de futebol).
A orientacao mandante/visitante segue o Controle; o job de email reorienta
o placar real da API para esses lados, entao um eventual flip nao afeta a conta.
"""
from __future__ import annotations

import csv

import openpyxl

FILE = "Controle_Boloes_Copa2026.xlsx"
SHEET = "Fase de Grupos"
OUT = "bolao_polla/palpites.csv"

# Codigo PT (Controle) -> nome canonico em ingles.
CODE_TO_EN = {
    "MEX": "Mexico", "COR": "South Korea", "AFS": "South Africa", "CZE": "Czechia",
    "CAN": "Canada", "SUI": "Switzerland", "CAT": "Qatar", "BOS": "Bosnia and Herzegovina",
    "BRA": "Brazil", "MAR": "Morocco", "ESC": "Scotland", "HAI": "Haiti",
    "EUA": "United States", "AUS": "Australia", "PAR": "Paraguay", "TUR": "Turkey",
    "ALE": "Germany", "EQU": "Ecuador", "CDM": "Ivory Coast", "CUR": "Curacao",
    "HOL": "Netherlands", "JAP": "Japan", "TUN": "Tunisia", "SUE": "Sweden",
    "BEL": "Belgium", "IRA": "Iran", "EGT": "Egypt", "NZL": "New Zealand",
    "ESP": "Spain", "URU": "Uruguay", "SAU": "Saudi Arabia", "CPV": "Cape Verde",
    "FRA": "France", "SEN": "Senegal", "NOR": "Norway", "IRQ": "Iraq",
    "ARG": "Argentina", "AUT": "Austria", "AGL": "Algeria", "JRD": "Jordan",
    "POR": "Portugal", "COL": "Colombia", "UZB": "Uzbekistan", "RDC": "DR Congo",
    "ING": "England", "CRO": "Croatia", "PAN": "Panama", "GAN": "Ghana",
}

# Composicao oficial dos grupos (sorteio real 2026) para verificar o mapa.
OFFICIAL_GROUPS = {
    "A": {"Mexico", "South Korea", "South Africa", "Czechia"},
    "B": {"Canada", "Switzerland", "Qatar", "Bosnia and Herzegovina"},
    "C": {"Brazil", "Morocco", "Scotland", "Haiti"},
    "D": {"United States", "Australia", "Paraguay", "Turkey"},
    "E": {"Germany", "Ecuador", "Ivory Coast", "Curacao"},
    "F": {"Netherlands", "Japan", "Tunisia", "Sweden"},
    "G": {"Belgium", "Iran", "Egypt", "New Zealand"},
    "H": {"Spain", "Uruguay", "Saudi Arabia", "Cape Verde"},
    "I": {"France", "Senegal", "Norway", "Iraq"},
    "J": {"Argentina", "Austria", "Algeria", "Jordan"},
    "K": {"Portugal", "Colombia", "Uzbekistan", "DR Congo"},
    "L": {"England", "Croatia", "Panama", "Ghana"},
}


def _to_int(v):
    return int(v) if isinstance(v, (int, float)) else None


def export() -> int:
    wb = openpyxl.load_workbook(FILE, data_only=True)
    ws = wb[SHEET]
    rows = []
    seen_by_group: dict[str, set] = {}
    for r in range(4, ws.max_row + 1):
        jogo = ws.cell(r, 2).value
        if not isinstance(jogo, int):
            continue
        grupo = ws.cell(r, 1).value
        hc, ac = ws.cell(r, 3).value, ws.cell(r, 4).value
        home, away = CODE_TO_EN.get(hc), CODE_TO_EN.get(ac)
        if home is None or away is None:
            raise ValueError(f"Codigo sem mapeamento na linha {r}: {hc}, {ac}")
        seen_by_group.setdefault(grupo, set()).update({home, away})
        rows.append({
            "grupo": grupo, "jogo": jogo, "home": home, "away": away,
            "zap_h": _to_int(ws.cell(r, 5).value), "zap_a": _to_int(ws.cell(r, 6).value),
            "cli_h": _to_int(ws.cell(r, 7).value), "cli_a": _to_int(ws.cell(r, 8).value),
            "polla_h": _to_int(ws.cell(r, 9).value), "polla_a": _to_int(ws.cell(r, 10).value),
        })

    # Verificacao: cada grupo deve conter exatamente os 4 times oficiais.
    for g, teams in seen_by_group.items():
        expected = OFFICIAL_GROUPS.get(g)
        if expected and teams != expected:
            raise ValueError(f"Grupo {g} divergente. CSV={teams} oficial={expected}")

    with open(OUT, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)
    return len(rows)


if __name__ == "__main__":
    n = export()
    print(f"Exportados {n} jogos para {OUT}")
