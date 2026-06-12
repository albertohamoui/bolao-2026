"""
Insere formulas de pontuacao automatica no Controle_Boloes_Copa2026.xlsx.

Para cada jogo da 'Fase de Grupos', preenche as colunas:
    M (Pts Zap), N (Pts Cli), O (Pts Polla)
com formulas que comparam o palpite de cada bolao (M/V) ao resultado real
(K/L) segundo as regras de cada regulamento. Enquanto o resultado real
estiver em branco, a celula fica vazia.

Colunas:
    E,F = Zap (M,V)   G,H = Cli (M,V)   I,J = Polla (M,V)
    K,L = Real (M,V)  M,N,O = Pts Zap/Cli/Polla

Regras (fase de grupos):
    Zap : exato 10 | dif 6 | gols vencedor 5 | gols perdedor 4 | so vencedor 3
          empate exato 10 | empate placar errado 6 | resto 0
    Cli : exato 5 (+1 se 5 gols totais, +2 se >=6) | empate certo errado 3
          vitoria certa errado 2 | resto 0
    Polla: exato 3 | vencedor/empate certo 1 | resto 0
"""
from __future__ import annotations

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

FILE = "Controle_Boloes_Copa2026.xlsx"
SHEET = "Fase de Grupos"


def zap_formula(r: int) -> str:
    return (
        f'=IF(OR($K{r}="",$L{r}="",$E{r}="",$F{r}=""),"",'
        f'IF($E{r}=$F{r},'
        f'IF($K{r}=$L{r},IF($E{r}=$K{r},10,6),0),'
        f'IF(OR($K{r}=$L{r},SIGN($E{r}-$F{r})<>SIGN($K{r}-$L{r})),0,'
        f'IF(AND($E{r}=$K{r},$F{r}=$L{r}),10,'
        f'IF(($E{r}-$F{r})=($K{r}-$L{r}),6,'
        f'IF(MAX($E{r},$F{r})=MAX($K{r},$L{r}),5,'
        f'IF(MIN($E{r},$F{r})=MIN($K{r},$L{r}),4,3)))))))'
    )


def cli_formula(r: int) -> str:
    return (
        f'=IF(OR($K{r}="",$L{r}="",$G{r}="",$H{r}=""),"",'
        f'IF(AND($G{r}=$K{r},$H{r}=$L{r}),'
        f'5+IF(($K{r}+$L{r})>=6,2,IF(($K{r}+$L{r})=5,1,0)),'
        f'IF(AND($G{r}=$H{r},$K{r}=$L{r}),3,'
        f'IF(AND($G{r}<>$H{r},$K{r}<>$L{r},SIGN($G{r}-$H{r})=SIGN($K{r}-$L{r})),2,0))))'
    )


def polla_formula(r: int) -> str:
    return (
        f'=IF(OR($K{r}="",$L{r}="",$I{r}="",$J{r}=""),"",'
        f'IF(AND($I{r}=$K{r},$J{r}=$L{r}),3,'
        f'IF(SIGN($I{r}-$J{r})=SIGN($K{r}-$L{r}),1,0)))'
    )


def add_formulas(wb: openpyxl.Workbook) -> int:
    ws = wb[SHEET]
    center = Alignment(horizontal="center")
    n = 0
    for r in range(1, ws.max_row + 1):
        jogo = ws.cell(row=r, column=2).value  # coluna B
        if not isinstance(jogo, int):
            continue  # pula cabecalho e banners de grupo
        ws.cell(row=r, column=13).value = zap_formula(r)    # M
        ws.cell(row=r, column=14).value = cli_formula(r)    # N
        ws.cell(row=r, column=15).value = polla_formula(r)  # O
        for col in (13, 14, 15):
            ws.cell(row=r, column=col).alignment = center
        n += 1
    return n


def add_totais(wb: openpyxl.Workbook) -> None:
    if "Totais" in wb.sheetnames:
        del wb["Totais"]
    ws = wb.create_sheet("Totais", index=1)

    title_fill = PatternFill("solid", fgColor="1F4E78")
    head_fill = PatternFill("solid", fgColor="BDD7EE")
    white_bold = Font(bold=True, color="FFFFFF", size=12)
    bold = Font(bold=True)
    center = Alignment(horizontal="center")

    ws["A1"] = "Placar de Pontos — Copa 2026"
    ws["A1"].font = white_bold
    ws["A1"].fill = title_fill
    ws.merge_cells("A1:C1")
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["Bolão", "Fase de Grupos", "Jogos pontuados"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = bold
        cell.fill = head_fill
        cell.alignment = center

    rows = [
        ("Bolão do Zap", "M"),
        ("Bolão do Cli", "N"),
        ("Polla 26", "O"),
    ]
    for i, (nome, col) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=nome).font = bold
        ws.cell(row=i, column=2,
                value=f"=SUM('{SHEET}'!{col}:{col})").alignment = center
        ws.cell(row=i, column=3,
                value=f"=COUNT('{SHEET}'!{col}:{col})").alignment = center

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16


def update_legenda(wb: openpyxl.Workbook) -> None:
    ws = wb["Legenda"]
    notas = [
        ("", ""),
        ("Pontuacao automatica", "Colunas M/N/O calculam sozinhas quando voce preenche Real (K/L)."),
        ("Bolao do Zap (grupos)",
         "Exato 10 | dif de gols 6 | gols do vencedor 5 | gols do perdedor 4 | "
         "so vencedor 3 | empate exato 10 | empate placar errado 6 | resto 0."),
        ("Bolao do Cli (grupos)",
         "Exato 5 (+1 se 5 gols no total, +2 se 6+ gols) | empate certo placar "
         "errado 3 | vitoria certa placar errado 2 | resto 0."),
        ("Polla 26",
         "Placar exato 3 | so o vencedor/empate certo 1 | resto 0."),
        ("Mata-mata",
         "Ainda nao consolidado. Zap dobra a tabela (exato 15, dif 9, gols venc 8, "
         "gols perd 6, so venc 5). Cli e Polla mantem os mesmos valores dos grupos."),
    ]
    start = ws.max_row + 1
    for i, (a, b) in enumerate(notas, start=start):
        ws.cell(row=i, column=1, value=a).font = Font(bold=bool(a))
        ws.cell(row=i, column=2, value=b)


def main() -> None:
    wb = openpyxl.load_workbook(FILE)
    n = add_formulas(wb)
    add_totais(wb)
    update_legenda(wb)
    wb.save(FILE)
    print(f"Formulas inseridas em {n} jogos. Aba 'Totais' e Legenda atualizadas.")
    print(f"Arquivo salvo: {FILE}")


if __name__ == "__main__":
    main()
