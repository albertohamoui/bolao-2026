"""
Gera os palpites de La Polla 2026 (polla26.com) maximizando a pontuacao.

Pontuacao da Polla (Reglas de La Polla 2026):
    - Placar exato (vitoria ou empate): 3 pontos
    - So acertou vencedor/empate (placar errado): 1 ponto
    - Errou o resultado: 0 pontos

Isso DIFERE do bolao do Zap/Cli (7 faixas). Logo o placar otimo muda:
sob a regra 3/1, o EV de um palpite (a,b) e:

    EV(a,b) = P(mesmo resultado que (a,b)) + 2 * P(placar exato (a,b))

O palpite Polla-otimo e o argmax de EV: o vencedor mais provavel + o placar
mais provavel dentro daquele resultado.

Motor de probabilidade: o MESMO modelo Dixon-Coles "smart baseline" ja definido.
Reaproveitamos os lambdas (lH, lA) de cada um dos 104 jogos do gabarito do Cli
(calibracao final: DC puro, xi=0.0005, qf=ON, cutoff=2020) e o rho do modelo,
reconstruindo a matriz de placares e aplicando o scoring da Polla.
"""
from __future__ import annotations

import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from scipy.stats import poisson

GABARITO_CLI = "bolao_cli/gabarito_bolao_cli.xlsx"
OUT_XLSX = "bolao_polla/Palpites_La_Polla_2026.xlsx"
RHO = -0.11978311298512935  # bolao2026_results.json -> model_params.rho
MAX_GOALS = 8


def tau(x: int, y: int, lx: float, my: float, rho: float) -> float:
    """Correcao Dixon-Coles para placares baixos."""
    if x == 0 and y == 0:
        return 1 - lx * my * rho
    if x == 0 and y == 1:
        return 1 + lx * rho
    if x == 1 and y == 0:
        return 1 + my * rho
    if x == 1 and y == 1:
        return 1 - rho
    return 1.0


def score_matrix(lh: float, la: float, rho: float = RHO) -> np.ndarray:
    """Matriz P(home=i, away=j) via Dixon-Coles, normalizada."""
    m = np.zeros((MAX_GOALS, MAX_GOALS))
    for i in range(MAX_GOALS):
        for j in range(MAX_GOALS):
            m[i, j] = max(
                poisson.pmf(i, lh) * poisson.pmf(j, la) * tau(i, j, lh, la, rho),
                0.0,
            )
    total = m.sum()
    return m / total if total > 0 else m


def polla_points(bet: tuple[int, int], actual: tuple[int, int]) -> int:
    """3 placar exato, 1 vencedor/empate certo, 0 caso contrario."""
    bh, ba = bet
    rh, ra = actual
    if (bh, ba) == (rh, ra):
        return 3
    bet_sign = (bh > ba) - (bh < ba)
    res_sign = (rh > ra) - (rh < ra)
    return 1 if bet_sign == res_sign else 0


def polla_optimal(m: np.ndarray, max_bet_goals: int = 6) -> tuple[tuple[int, int], float]:
    """Placar que maximiza o EV de pontos da Polla."""
    best_score, best_ev = (0, 0), -1.0
    n = m.shape[0]
    for a in range(max_bet_goals + 1):
        for b in range(max_bet_goals + 1):
            ev = 0.0
            for i in range(n):
                for j in range(n):
                    p = m[i, j]
                    if p > 0:
                        ev += p * polla_points((a, b), (i, j))
            if ev > best_ev:
                best_ev, best_score = ev, (a, b)
    return best_score, best_ev


def result_probs(m: np.ndarray) -> tuple[float, float, float]:
    """(P home, P draw, P away)."""
    p_home = float(np.sum(np.tril(m, -1)))
    p_draw = float(np.sum(np.diag(m)))
    p_away = float(np.sum(np.triu(m, 1)))
    return p_home, p_draw, p_away


def load_matches() -> tuple[list[dict], list[dict]]:
    """Le os 104 jogos do gabarito do Cli (grupos + mata-mata) com lH, lA."""
    wb = openpyxl.load_workbook(GABARITO_CLI)
    ws = wb["Bolao do Cli"]
    groups, knockout = [], []
    for row in ws.iter_rows(min_row=1, values_only=True):
        num, fase, casa, fora, _palpite, _dec, lh, la, _pe = row
        if not (isinstance(casa, str) and isinstance(fora, str)):
            continue
        if not (isinstance(lh, (int, float)) and isinstance(la, (int, float))):
            continue
        rec = {"fase": fase, "home": casa, "away": fora,
               "lh": float(lh), "la": float(la), "pe_cli": _pe}
        if isinstance(fase, str) and fase.startswith("Grupo"):
            groups.append(rec)
        elif fase in ("R32", "R16", "QF", "SF", "3o lugar", "FINAL"):
            knockout.append(rec)
    return groups, knockout


def build_rows(matches: list[dict]) -> list[dict]:
    out = []
    for rec in matches:
        m = score_matrix(rec["lh"], rec["la"])
        (a, b), ev = polla_optimal(m)
        p_home, p_draw, p_away = result_probs(m)
        # vencedor previsto / quem avanca (empate -> favorito por lambda)
        if a > b:
            winner = rec["home"]
        elif b > a:
            winner = rec["away"]
        else:
            winner = rec["home"] if rec["lh"] >= rec["la"] else rec["away"]
        out.append({
            **rec,
            "palpite": f"{a}x{b}",
            "ev": ev,
            "p_home": p_home, "p_draw": p_draw, "p_away": p_away,
            "winner": winner,
            "p_exact": float(m[a, b]),
        })
    return out


def write_excel(group_rows: list[dict], ko_rows: list[dict]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "La Polla 2026"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    phase_fill = PatternFill("solid", fgColor="BDD7EE")
    white_bold = Font(bold=True, color="FFFFFF")
    bold = Font(bold=True)
    center = Alignment(horizontal="center")

    cols = ["#", "Fase", "Casa", "Fora", "PALPITE", "Avanca",
            "P(Casa)", "P(Empate)", "P(Fora)", "P(placar)", "EV pts"]
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = white_bold
        cell.alignment = center

    n = 0

    def add_phase_banner(text: str) -> None:
        ws.append([text] + [None] * (len(cols) - 1))
        r = ws.max_row
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(cols))
        cell = ws.cell(row=r, column=1)
        cell.fill = phase_fill
        cell.font = bold
        cell.alignment = Alignment(horizontal="left")

    def add_rows(rows: list[dict], knockout: bool) -> None:
        nonlocal n
        for rec in rows:
            n += 1
            ws.append([
                n, rec["fase"], rec["home"], rec["away"], rec["palpite"],
                rec["winner"] if knockout else "",
                round(rec["p_home"] * 100, 1),
                round(rec["p_draw"] * 100, 1),
                round(rec["p_away"] * 100, 1),
                round(rec["p_exact"] * 100, 1),
                round(rec["ev"], 2),
            ])
            pal = ws.cell(row=ws.max_row, column=5)
            pal.font = bold
            pal.alignment = center

    add_phase_banner("FASE DE GRUPOS")
    add_rows(group_rows, knockout=False)
    add_phase_banner("FASE ELIMINATORIA (bracket simulado pelo modelo)")
    add_rows(ko_rows, knockout=True)

    widths = [5, 22, 22, 22, 9, 16, 9, 11, 9, 11, 8]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    wb.save(OUT_XLSX)


def main() -> None:
    groups, knockout = load_matches()
    print(f"Jogos lidos: {len(groups)} grupos + {len(knockout)} mata-mata")

    group_rows = build_rows(groups)
    ko_rows = build_rows(knockout)

    # Sanity: P(empate) reconstruido deve bater com o gabarito do Cli
    print("\nValidacao P(empate) reconstruido vs gabarito Cli (5 amostras):")
    for rec in group_rows[:5]:
        print(f"  {rec['home']:<14} x {rec['away']:<14} "
              f"recon={rec['p_draw']*100:4.0f}%  cli={rec['pe_cli']}")

    total_ev = sum(r["ev"] for r in group_rows + ko_rows)
    print(f"\nEV total esperado (Polla): {total_ev:.1f} pts em "
          f"{len(group_rows)+len(ko_rows)} jogos")

    write_excel(group_rows, ko_rows)
    print(f"\nExcel salvo: {OUT_XLSX}")


if __name__ == "__main__":
    main()
